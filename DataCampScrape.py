from selenium import webdriver
from bs4 import BeautifulSoup
import re
import pandas as pd
import datetime
from openpyxl import load_workbook

# webdriver is needed in order to run javascript on webpage
driver = webdriver.Chrome('C:\ProgramData\chromedriver')


# Create DataFrames for data previously pulled
existingExercises = pd.read_excel('C:\\Users\\christian.taylor\\OneDrive - Decisive Data\\DataCampLeaderboard\\DataCamp.xlsx', sheetname = 'Exercises')
existingCourses = pd.read_excel('C:\\Users\\christian.taylor\\OneDrive - Decisive Data\\DataCampLeaderboard\\DataCamp.xlsx', sheetname = 'Courses')
existingTracks = pd.read_excel('C:\\Users\\christian.taylor\\OneDrive - Decisive Data\\DataCampLeaderboard\\DataCamp.xlsx', sheetname = 'Tracks')
existingXP = pd.read_excel('C:\\Users\\christian.taylor\\OneDrive - Decisive Data\\DataCampLeaderboard\\DataCamp.xlsx', sheetname = 'Topics')


# Create a list of user names
users = list(pd.read_csv('C:\\Users\\christian.taylor\\OneDrive - Decisive Data\\DataCampLeaderboard\\DataCampUsers.csv', header=None)[0])


# initialize lists & dictionaries
exercisesAced_list = []
coursesCompleted_dict = {}
tracksCompleted_dict = {}
topicXP_dict = {}

# loop through each user, visiting their Data Camp public profile and scraping the data
for userName in users:
    url = 'https://www.datacamp.com/profile/' + userName
    
    # We can't run these commented requests lines because we need to run the javascript on the page
    #import requests
    #r = requests.get(url, allow_redirects=False)
    #html = r.text
    
    # Get the HTML and store it as a BeautifulSoup object
    driver.get(url)
    htmlSource = driver.page_source
    soup = BeautifulSoup(htmlSource, "lxml")
    
    
    # find the number of exercises aced and append to list
    exercisesAced = int(soup.find_all("strong", { "class" : "stats-block__number" })[-1].text)
    exercisesAced_list.append(exercisesAced)
    
    # find each users XP for each topic and add the topicXP_dict
    userTopics = soup.find_all("div", {"class" : "topic-block__content"})
    userTopics_Names = []
    userTopics_XP = []
    for topic in userTopics:
        if not topic.find("h5") == -1:
            topicName = topic.find("h4").contents[0]
            topicXP = int(re.match(r'[0-9]*', topic.find("p").contents[0]).group())
            userTopics_Names.append(topicName)
            userTopics_XP.append(topicXP)       
    topicXP_dict[userName] = [userTopics_Names,userTopics_XP]
    
    
    # find all tracks the user has completed and add them the the tracks dictionary
    userTracks = [c.contents[0] for c in soup.find_all("h4", {"class" : "track-block__title"})]
    tracksCompleted_dict[userName] = userTracks
    
    # find all courses the user has completed and add them the the courses dictionary
    userCourses = [c.contents[0] for c in soup.find_all("h4", {"class" : "course-block__title"})]
    coursesCompleted_dict[userName] = userCourses

#End of for loop


# Create DataFrame for Exercises Aced data
exercises_df = pd.DataFrame({'UserName':users,'ExercisesAced':exercisesAced_list})

# Use existingExercises and exercises_df to find the new Exercises Aced
exerciseSum = existingExercises.groupby(['UserName'], as_index=False)[['ExercisesAced']].sum()
exerciseJoin = pd.merge(exercises_df,exerciseSum,'left',on='UserName')
exerciseJoin['ExercisesAced'] = exerciseJoin['ExercisesAced_x'] - exerciseJoin['ExercisesAced_y']
exerciseNew = exerciseJoin[exerciseJoin['ExercisesAced'] != 0]
exerciseNew = exerciseNew[['ExercisesAced', 'UserName']]
exerciseNew['Date'] = datetime.datetime.now().date()



# Create DataFrame for Courses Completed data
userCourses_list = []
for k, v in coursesCompleted_dict.items():
    for i in v:
        userCourses_list.append([k,i])
userCourses_df = pd.DataFrame(userCourses_list,columns=['UserName','CourseName'])

# Use existingCourses and userCourses_df to find the new Courses Completed
coursesJoin = pd.merge(userCourses_df,existingCourses,'outer',on=['UserName','CourseName'])
coursesNew = coursesJoin[coursesJoin['DateCompleted'].isnull()]
coursesNew = coursesNew[['UserName','CourseName']]
coursesNew['DateCompleted'] = datetime.datetime.now().date()


# Create DataFrame for Tracks Completed data
userTracks_list = []
for k, v in tracksCompleted_dict.items():
    for i in v:
        userTracks_list.append([k,i])
userTracks_df = pd.DataFrame(userTracks_list,columns=['UserName','TrackName'])

# Use existingTracks and userTracks_df to find the new Tracks Completed
tracksJoin = pd.merge(userTracks_df,existingTracks,'outer',on=['UserName','TrackName'])
tracksNew = tracksJoin[tracksJoin['DateCompleted'].isnull()]
tracksNew = tracksNew[['UserName','TrackName']]
tracksNew['DateCompleted'] = datetime.datetime.now().date()


# Create DataFrame for Topic XP data
userTopics_list = []
userTopicsNames_list = []
userTopicsXP_list = []
for k, v in topicXP_dict.items():
    for i in v[0]:
        userTopics_list.append(k)
        userTopicsNames_list.append(i)
    for i in v[1]:
        userTopicsXP_list.append(i)
topicXP_df = pd.DataFrame(list(zip(userTopics_list,userTopicsNames_list,userTopicsXP_list)),columns=['UserName','TopicName','XP'])

# Use existingXP and topicXP_df to find the new XP per Topic
topicXPSum = existingXP.groupby(['UserName', 'TopicName'], as_index=False)[['XP']].sum()
topicXPJoin = pd.merge(topicXP_df,topicXPSum,'left',on=['UserName','TopicName'])
topicXPJoin['XP'] = topicXPJoin['XP_x'] - topicXPJoin['XP_y']
topicXPNew = topicXPJoin[topicXPJoin['XP'] != 0]
topicXPNew = topicXPNew[['UserName','TopicName','XP']]
topicXPNew['Date'] = datetime.datetime.now().date()




# load excel workbook
book = load_workbook('C:\\Users\\christian.taylor\\OneDrive - Decisive Data\\DataCampLeaderboard\DataCamp.xlsx')
writer = pd.ExcelWriter('C:\\Users\\christian.taylor\\OneDrive - Decisive Data\\DataCampLeaderboard\DataCamp.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)


# append new exercises, courses, tracks, and XP to existing sheets
exerciseNew.to_excel(writer, "Exercises", index=False, header = False, startrow = writer.sheets['Exercises'].max_row)
coursesNew.to_excel(writer, "Courses", index=False, header = False, startrow = writer.sheets['Courses'].max_row)
tracksNew.to_excel(writer, "Tracks", index=False, header = False, startrow = writer.sheets['Tracks'].max_row)
topicXPNew.to_excel(writer, "Topics", index=False, header = False, startrow = writer.sheets['Topics'].max_row)


# save file  
writer.save()

driver.quit()






